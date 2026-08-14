# ==========================================================
# GLASS PI VERIFICATION SYSTEM
# REPORT GENERATOR MODULE
# Enterprise Edition
# ==========================================================

import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

# ==========================================================
# REPORT BUILDER FUNCTION
# ==========================================================

def build_report(result_df: pd.DataFrame, report_path: str) -> None:
    """
    Generates a professionally styled Excel verification report
    with color-coded statuses.
    """
    logger.info("Building Enterprise Excel Report -> %s", report_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Verification Summary"

    # Ensure gridlines are visible
    ws.views.sheetView[0].showGridLines = True

    # Header Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Write Headers
    headers = list(result_df.columns)
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border_thin

    ws.row_dimensions[1].height = 25

    # Status Fills
    match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Soft Green
    mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Soft Red
    notfound_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Soft Yellow

    match_font = Font(name="Calibri", size=10, color="006100", bold=True)
    mismatch_font = Font(name="Calibri", size=10, color="9C0006", bold=True)
    notfound_font = Font(name="Calibri", size=10, color="9C6500", bold=True)

    # Write Data Rows
    for row_idx, row_data in enumerate(result_df.itertuples(index=False), start=2):
        ws.append(list(row_data))
        ws.row_dimensions[row_idx].height = 20

        result_val = getattr(row_data, "Result", "")

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border_thin
            cell.alignment = Alignment(vertical="center", horizontal="left" if col_idx == 1 else "center")
            cell.font = Font(name="Calibri", size=10)

            # Apply conditional styling on 'Result' column
            if headers[col_idx - 1] == "Result":
                if result_val == "MATCHING":
                    cell.fill = match_fill
                    cell.font = match_font
                elif result_val == "MISMATCH":
                    cell.fill = mismatch_fill
                    cell.font = mismatch_font
                elif result_val == "NOT FOUND":
                    cell.fill = notfound_fill
                    cell.font = notfound_font

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(report_path)
    logger.info("Report successfully generated and saved.")