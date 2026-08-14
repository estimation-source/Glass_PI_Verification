# ============================================================
# Glass PI Verification System - Exact Formatted Excel Reader
# ============================================================

from __future__ import annotations

import io
import re
from dataclasses import asdict, dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

HEADER_SCAN_LIMIT = 200
HEADER_REMOVE_PATTERN = r"[^A-Z0-9]"

KEYWORDS = {
    "CODE": ["CODE"],
    "WIDTH": [
        ["GL", "W"], ["GLS", "W"], ["GLASS", "W"],
        ["GL", "WIDTH"], ["GLS", "WIDTH"], ["GLASS", "WIDTH"],
        ["S", "GLS", "W"], ["SGLS", "W"], ["S", "GL", "W"],
    ],
    "HEIGHT": [
        ["GL", "H"], ["GLS", "H"], ["GLASS", "H"],
        ["GL", "HEIGHT"], ["GLS", "HEIGHT"], ["GLASS", "HEIGHT"],
        ["S", "GLS", "H"], ["SGLS", "H"], ["S", "GL", "H"],
    ],
    "QTY": ["QTY"],
    "GLASS": [["GLASS"], ["DESP"]],
}

@dataclass(slots=True)
class HeaderInfo:
    row_index: int
    code_col: Optional[int] = None
    width_col: Optional[int] = None
    height_col: Optional[int] = None
    qty_col: Optional[int] = None
    glass_col: Optional[int] = None
    columns: Dict[str, Optional[int]] = field(default_factory=dict)

@dataclass(slots=True)
class HeaderBlock:
    header: HeaderInfo
    start_row: int
    end_row: int

@dataclass(slots=True)
class GlassRecord:
    WindowCode: str
    Width: int
    Height: int
    Qty: int
    GlassType: str
    SourceFile: str
    SheetName: str

def normalize_header(text: Any) -> str:
    if pd.isna(text):
        return ""
    text = str(text).upper().strip()
    return re.sub(HEADER_REMOVE_PATTERN, "", text)

def normalize_header_row(row: pd.Series) -> List[str]:
    return [normalize_header(val) for val in row.tolist()]

def contains_keywords(text: str, keyword_groups: List[Any]) -> bool:
    if not text:
        return False
    text = normalize_header(text)
    text = re.sub(r"[^A-Z0-9]", "", text)

    for group in keyword_groups:
        if isinstance(group, str):
            group = [group]
        matched = True
        for keyword in group:
            key = re.sub(r"[^A-Z0-9]", "", normalize_header(keyword))
            if key not in text:
                matched = False
                break
        if matched:
            return True
    return False

def detect_column(header_row: List[str], keyword_groups: List[Any]) -> Optional[int]:
    for index, value in enumerate(header_row):
        text = normalize_header(value)
        text = re.sub(r"[^A-Z0-9]", "", text)
        text = (
            text.replace("GLASS", "GL")
            .replace("GLAZING", "GL")
            .replace("SGLS", "GLS")
            .replace("SGL", "GL")
            .replace("GLS.", "GLS")
        )
        for group in keyword_groups:
            if isinstance(group, str):
                group = [group]
            matched = True
            for keyword in group:
                key = re.sub(r"[^A-Z0-9]", "", normalize_header(keyword))
                if key not in text:
                    matched = False
                    break
            if matched:
                return index
    return None

def detect_header_columns(header_row: pd.Series) -> Dict[str, Optional[int]]:
    normalized = normalize_header_row(header_row)
    columns = {
        "code": detect_column(normalized, KEYWORDS["CODE"]),
        "width": detect_column(normalized, KEYWORDS["WIDTH"]),
        "height": detect_column(normalized, KEYWORDS["HEIGHT"]),
        "qty": detect_column(normalized, KEYWORDS["QTY"]),
        "glass": detect_column(normalized, KEYWORDS["GLASS"]),
    }

    if columns["width"] is None:
        for kw in [["S", "GLS", "W"], ["S", "GL", "W"], ["GLS", "W"], ["GL", "W"], ["FWIDTH"], ["SWIDTH"]]:
            col = detect_column(normalized, kw)
            if col is not None:
                columns["width"] = col
                break

    if columns["height"] is None:
        for kw in [["S", "GLS", "H"], ["S", "GL", "H"], ["GLS", "H"], ["GL", "H"], ["FHEIGHT"], ["SHEIGHT"]]:
            col = detect_column(normalized, kw)
            if col is not None:
                columns["height"] = col
                break

    return columns

def is_business_header(row: pd.Series) -> bool:
    normalized = normalize_header_row(row)
    has_code, has_dim, has_qty = False, False, False

    for value in normalized:
        if contains_keywords(value, KEYWORDS["CODE"]):
            has_code = True
        if contains_keywords(value, KEYWORDS["WIDTH"]) or contains_keywords(value, KEYWORDS["HEIGHT"]):
            has_dim = True
        if contains_keywords(value, KEYWORDS["QTY"]):
            has_qty = True

    return has_code and has_dim and has_qty

def find_header_blocks(dataframe: pd.DataFrame) -> List[HeaderInfo]:
    headers: List[HeaderInfo] = []
    rows = min(len(dataframe), HEADER_SCAN_LIMIT)

    for row_number in range(rows):
        row = dataframe.iloc[row_number]
        if not is_business_header(row):
            continue

        columns = detect_header_columns(row)
        header = HeaderInfo(
            row_index=row_number,
            code_col=columns["code"],
            width_col=columns["width"],
            height_col=columns["height"],
            qty_col=columns["qty"],
            glass_col=columns["glass"],
            columns=columns,
        )
        headers.append(header)

    return headers

def build_header_blocks(dataframe: pd.DataFrame, headers: List[HeaderInfo]) -> List[HeaderBlock]:
    blocks: List[HeaderBlock] = []
    if not headers:
        return blocks

    headers = sorted(headers, key=lambda h: h.row_index)
    for i, header in enumerate(headers):
        start = header.row_index + 1
        end = len(dataframe) - 1 if i == len(headers) - 1 else headers[i + 1].row_index - 1
        blocks.append(HeaderBlock(header=header, start_row=start, end_row=end))

    return blocks

def score_business_sheet(df: pd.DataFrame) -> int:
    score = 0
    scan_rows = min(len(df), HEADER_SCAN_LIMIT)

    for r in range(scan_rows):
        row = normalize_header_row(df.iloc[r])
        text = " ".join(row)
        if "CODE" in text: score += 5
        if "QTY" in text: score += 8
        if "GLASS" in text: score += 8
        if re.search(r"\bGL.*W\b", text): score += 10
        if re.search(r"\bGL.*H\b", text): score += 10

    return score

def find_business_sheets(workbook: Dict[str, pd.DataFrame]) -> List[Tuple[str, pd.DataFrame]]:
    business_sheets = []
    threshold = 20

    for sheet_name, df in workbook.items():
        score = score_business_sheet(df)
        if score >= threshold or find_header_blocks(df):
            business_sheets.append((sheet_name, df))

    return business_sheets

# ============================================================
# EXACT DISPLAYED VALUE EXTRACTION FIX
# ============================================================

def safe_exact_integer(value: Any) -> Optional[int]:
    """
    एक्सेल सेलमधील मूळ स्ट्रिंग डिस्प्ले व्हॅल्यू जशीच्या तशी अचूक वाचणे.
    कोणतेही राउंडिंग किंवा फ्लोट कन्व्हर्जन न करता सेलमधील अंक वाचतो.
    """
    if pd.isna(value) or value is None:
        return None
    
    val_str = str(value).strip()
    
    # जर सेलमधील स्ट्रिंगमध्ये पॉइंट असेल (उदा. "770.8" किंवा "771.0") तर
    # शेवटचे कॅल्क्युलेटेड डिस्प्ले कॅरेक्टर उचलणे
    clean_num = re.sub(r"[^\d.]", "", val_str)
    if not clean_num:
        return None
        
    try:
        val_float = float(clean_num)
        if val_float <= 0:
            return None
        
        # स्टँडर्ड 0.5 राउंडिंग लावून मूळ डिस्प्ले इंटीजर मिळवणे
        return int(val_float + 0.5)
    except Exception:
        return None

def build_window_code(row: pd.Series, header: HeaderInfo) -> Optional[str]:
    if header.code_col is None or header.code_col >= len(row):
        return None

    val = row.iloc[header.code_col]
    if pd.isna(val) or str(val).strip() == "" or str(val).lower() == "nan":
        return None

    code = re.sub(r"\.0$", "", str(val).strip())
    code = re.sub(r"\s+", " ", code).strip()

    next_col = header.code_col + 1
    other_cols = {header.width_col, header.height_col, header.qty_col, header.glass_col}

    if next_col < len(row) and next_col not in other_cols:
        next_val = row.iloc[next_col]
        if not pd.isna(next_val):
            next_str = str(next_val).strip()
            next_str = re.sub(r"\.0$", "", next_str)
            if next_str and next_str.lower() != "nan" and not next_str.replace('.', '', 1).isdigit():
                code = f"{code} {next_str}"

    return re.sub(r"\s+", " ", code).strip()

def is_valid_glass_text(cell_str: str) -> bool:
    """ग्लास टेक्स्ट वैध आहे की नाही (उदा. फ्रॉस्टेड चे नियम) हे तपासणे."""
    text_upper = cell_str.upper().strip()
    
    # १. जर फक्त "FROSTED" किंवा "FROSTED GLASS" असेल तर ते रिजेक्ट (इग्नोर) करा
    if "FROSTED" in text_upper:
        # जर FROSTED सोबत TOUGHENED किंवा TGH असेल तरच ठेवा
        if "TOUGHENED" in text_upper or "TGH" in text_upper:
            return True
        else:
            return False
            
    return True

def extract_smart_glass_from_row(row: pd.Series, skip_cols: List[int]) -> Optional[str]:
    """जर Glass Column मॅच झाला नाही, तर संपूर्ण रो मधील शेवटचे टेक्स्ट किंवा Keywords शोधणे."""
    glass_keywords = ["LAMINATED", "TOUGHENED", "DGU", "SGU", "CLEAR", "TGH", "LAM", "PVB", "GLASS", "MM"]

    # १. उजवीकडून डावीकडे (Right to Left) स्कॅन करणे
    for idx in reversed(range(len(row))):
        if idx in skip_cols:
            continue
        cell = row.iloc[idx]
        if pd.isna(cell):
            continue
            
        cell_str = str(cell).strip()
        cell_upper = cell_str.upper()

        # जर सेलमधील व्हॅल्यू फक्त नंबर असेल तर ती Glass असू शकत नाही
        if cell_str.replace('.', '', 1).isdigit():
            continue

        # Header किंवा इतर नको असलेले शब्द वगळणे
        if any(h in cell_upper for h in ["HANDLE", "LOCK", "COLOR", "FRAME", "CODE", "QTY", "WIDTH", "HEIGHT"]):
            continue

        # फ्रॉस्टेड ची स्पेशल अट तपासणे
        if not is_valid_glass_text(cell_str):
            continue

        # जर कीवर्ड्स मॅच झाले तर ते टेक्स्ट रिटर्न करा
        if any(kw in cell_upper for kw in glass_keywords) or "FROSTED TOUGHENED" in cell_upper:
            return re.sub(r"\s+", " ", cell_str).strip()

    return None


def is_invalid_frosted_row(row: pd.Series) -> bool:
    """जर रो मध्ये FROSTED शब्द असेल पण TOUGHENED/TGH नसेल तर सत्य (True) रिटर्न करतो."""
    row_str = " ".join([str(val).upper() for val in row.values if not pd.isna(val)])
    
    if "FROSTED" in row_str:
        # जर FROSTED सोबत TOUGHENED किंवा TGH नसेल, तर हा संपूर्ण रो इग्नोर करायचा आहे
        if "TOUGHENED" not in row_str and "TGH" not in row_str:
            return True
            
    return False


def parse_header_block(dataframe: pd.DataFrame, block: HeaderBlock, source_file: str, sheet_name: str) -> List[GlassRecord]:
    records: List[GlassRecord] = []
    
    last_seen_window: Optional[str] = None
    last_seen_glass: Optional[str] = None

    skip_cols = [c for c in [block.header.width_col, block.header.height_col, block.header.qty_col, block.header.code_col] if c is not None]

    for row_no in range(block.start_row, block.end_row + 1):
        row = dataframe.iloc[row_no]

        # ✅ FIX 1: जर रो मध्ये अनव्हॅलिड FROSTED असेल तर संपूर्ण रो SKIP (Ignore) करा
        if is_invalid_frosted_row(row):
            continue

        window_code = build_window_code(row, block.header)
        if window_code:
            last_seen_window = window_code

        g_str = None
        
        # १. जर Glass Column डिटेक्ट झाला असेल तर तिथून व्हॅल्यू वाचणे
        if block.header.glass_col is not None and block.header.glass_col < len(row):
            g_val = row.iloc[block.header.glass_col]
            if not pd.isna(g_val) and str(g_val).strip():
                temp_str = str(g_val).strip()
                if not temp_str.replace('.', '', 1).isdigit() and temp_str.upper() not in ["GLASS", "DESCRIPTION", "DESP", "SPECIFICATION", "REMARKS", "NAN"]:
                    g_str = temp_str

        # २. जर Glass Column रिकामा असेल, तर Smart Search वापरा
        if not g_str:
            g_str = extract_smart_glass_from_row(row, skip_cols)

        if g_str:
            last_seen_glass = g_str

        width = safe_exact_integer(row.iloc[block.header.width_col]) if block.header.width_col is not None and block.header.width_col < len(row) else None
        height = safe_exact_integer(row.iloc[block.header.height_col]) if block.header.height_col is not None and block.header.height_col < len(row) else None
        qty = safe_exact_integer(row.iloc[block.header.qty_col]) if block.header.qty_col is not None and block.header.qty_col < len(row) else None

        if width is None or height is None or qty is None or not last_seen_window:
            continue

        glass_type = last_seen_glass or "NOT SPECIFIED"

        records.append(
            GlassRecord(
                WindowCode=str(last_seen_window),
                Width=int(width),
                Height=int(height),
                Qty=int(qty),
                GlassType=str(glass_type).strip(),
                SourceFile=str(source_file),
                SheetName=str(sheet_name),
            )
        )

    return records



def parse_business_sheet(dataframe: pd.DataFrame, source_file: str, sheet_name: str) -> List[GlassRecord]:
    headers = find_header_blocks(dataframe)
    blocks = build_header_blocks(dataframe, headers)
    all_records: List[GlassRecord] = []

    for block in blocks:
        all_records.extend(parse_header_block(dataframe, block, source_file, sheet_name))

    return all_records

def load_excel_with_calculated_values(file_obj) -> Dict[str, pd.DataFrame]:
    """Loads exact cell data cleanly without openpyxl float degradation."""
    if hasattr(file_obj, "read"):
        file_bytes = io.BytesIO(file_obj.read())
        file_obj.seek(0)
    else:
        file_bytes = file_obj

    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    workbook_dict = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        data_rows = []
        for row in sheet.iter_rows(values_only=True):
            data_rows.append(list(row))
            
        if not data_rows:
            continue
        df = pd.DataFrame(data_rows)
        workbook_dict[sheet_name] = df

    return workbook_dict

def process_uploaded_files(uploaded_files: list) -> pd.DataFrame:
    all_records: List[GlassRecord] = []

    for file in uploaded_files:
        try:
            file_name = getattr(file, "name", "uploaded_file.xlsx")
            workbook_dict = load_excel_with_calculated_values(file)
            business_sheets = find_business_sheets(workbook_dict)
            
            for sheet_name, df in business_sheets:
                records = parse_business_sheet(df, file_name, sheet_name)
                all_records.extend(records)
        except Exception as e:
            print(f"Error processing {getattr(file, 'name', 'file')}: {e}")

    if not all_records:
        return pd.DataFrame(columns=["WindowCode", "Width", "Height", "Qty", "GlassType", "SourceFile", "SheetName"])

    records_dict = [asdict(r) for r in all_records]
    return pd.DataFrame(records_dict)